# A

import os
import openpyxl
import time

import nest_asyncio
nest_asyncio.apply()

from requests_html import AsyncHTMLSession
from bs4 import BeautifulSoup
import asyncio

# asession = AsyncHTMLSession()

from playwright.async_api import async_playwright

async def get_roof_space(addr_num,lat,lon):
    url = f"https://sunroof.withgoogle.com/building/{str(round(float(lat),4))}/{str(round(float(lon),4))}/#?f=buy"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.goto(url, timeout=30000)
        html = await page.content()
        await browser.close()
        soup = BeautifulSoup(html, "html.parser")

        inp = soup.find("input", attrs={"aria-label": True})
        if not inp: inp = soup.find("input", attrs={"placeholder": True})

        if inp:
            address = inp.get("aria-label") or inp.get("placeholder") # as there are 2 options to where the address could be found
            # print("Address:", address)
            # print(addr_num, address.split(" ")[0])
            if addr_num not in address.split(" ")[0]: return "None"
        else:
            print("Couldn’t find the address input in the page.")
            return "None"
        

        upper_li = soup.find("li", attrs={"ng-if": "$ctrl.derivedValues.getMaxArrayAreaSqft()"})
        if upper_li:
            data_div = upper_li.find("div", class_="panel-fact-text md-body")
            if data_div:
                val = data_div.get_text(strip=True).split()[0]
                # print(val)
                return(val)
            else:
                print("Inner data div not found.")
                return("Inner data div not found.")
        else:
            print("Upper div with specified ng-if not found.")
            return("Upper div with specified ng-if not found.")


# open excel file
# wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
# ws = wb.active

# code to test our function
# test_lat, test_lon = ws.cell(row=4, column=2).value, ws.cell(row=4, column=3).value
# test_addr = ws.cell(row=4, column=1).value
# test_data = data = asyncio.run(get_roof_space(test_addr.split(" ")[0],test_lat,test_lon))
# print(test_data)


# writing results to excel
# for row in range(2,211):
#     addr = ws.cell(row=row, column=1).value
#     # print(addr_num)
#     lat = ws.cell(row=row, column=2).value
#     long = ws.cell(row=row, column=3).value
#     data = asyncio.run(scrape_roof_space(addr.split(" ")[0],lat,long))
#     print(f"{addr} → {data}")
#     ws.cell(row=row, column=5, value=data)
#     wb.save('res-econ_RA_data.xlsx')
#     time.sleep(1)
# print("Done updating Excel file.")