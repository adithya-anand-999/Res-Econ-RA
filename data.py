# import all data collection functions
# detailed comments of how each data collection function works is found in the other code files. Each file is named after the data point it collects.
from coordinates import get_coordinates
from capacity_factor import get_capacity_factor
from roof_space import get_roof_space
from zillow_url import get_zillow_url
from zillow_data import get_zillow_data
from zillow_data import get_snapshot_id

# required libraries for function get_data
import openpyxl
import asyncio
import time

# this function calls all our data collection functions and writes the results to the excel file
def get_data(excel_path):
    # open excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    end_row = ws.max_row

    # calls our data collection functions for each row specified in the args
    for row in range(2,end_row+1): # first arg inclusive, second arg exclusive. Because second arg exclusive we should do end_row+1
        addr = ws.cell(row=row, column=1).value
        if not addr: continue
        print(f'Parsing row {row-1} with address "{addr}"')
        lat, lon = get_coordinates(addr)
        ws.cell(row=row, column=8, value=get_capacity_factor(lat,lon))
        ws.cell(row=row, column=9, value=asyncio.run(get_roof_space(addr.split(" ")[0],lat,lon)))
        url = get_zillow_url(addr)
        ws.cell(row=row, column=10, value=url)
        if url == 'None': continue
        zillow_data = get_zillow_data(get_snapshot_id(url))
        if not zillow_data: continue
        for i,point in enumerate(zillow_data.values()): ws.cell(row=row, column=11+i, value=point if point else "None")
        wb.save(excel_path)
        time.sleep(1)
    # data collection complete
    print("Done updating Excel file.")


get_data("./given_excel.xlsx") # invoke our function

