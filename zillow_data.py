# E

# required libraries we use.
import requests
from time import sleep
import openpyxl

from config import BRIGHT_DATA # API key for Bright Data APIs

# data points we are collecting
required_keys = ["zestimate", "taxAssessedValue", "taxAssessedYear", "dateSoldString", "livingArea", "yearBuilt", "lotAreaValue", "lotAreaUnits"]

# function which converts our url for an address to a unique snapshot id. More information about snapshot ID can be found <link>, it is a step in the data collection
# framework found for Bright Data. Bright Data provides a third party zillow API we are using to collect information. 
# TODO: write comments explaining how the code in the function works.
def get_snapshot_id(addr_url):
    # boilerplate code from Bright Data website 
    url = "https://api.brightdata.com/datasets/v3/trigger"
    headers = {
        "Authorization": BRIGHT_DATA,
        "Content-Type": "application/json",
    }
    params = {
        "dataset_id": "gd_lfqkr8wm13ixtbd8f5",
        "include_errors": "true",
    }
    data = [
        {"url":addr_url},
    ]
    response = requests.post(url, headers=headers, params=params, json=data)
    return(response.json()['snapshot_id'])  # snapshot_id is a pointer to a zillow data object for this address

# function parses the snapshot ID, producing a json object with all the data for the address. From this we collect the specific data points required.

# TODO: write comments explaining how the code in the function works.
def get_zillow_data(snapshotID):
    # boilerplate code from Bright Data
    url = f"https://api.brightdata.com/datasets/v3/snapshot/{snapshotID}"
    headers = {"Authorization": BRIGHT_DATA}
    data = requests.request('GET', url, headers=headers).json()

    
    retry_count = 0

    while 'status' in data and retry_count < 3:
        retry_count += 1
        print(f"Zillow Data Collection: Retrying {retry_count}, sleeping 30s…")
        sleep(30)
        data = requests.request('GET', url, headers=headers).json()

    if 'status' in data:
        print(f"Scraping failed for snapshot {snapshotID}")
        return None
    
    # debugging code for seeing complete json data object
    # print(data)
    # print("data found, pause for debugging")
    # sleep(30)

    payload = {}
    for key in required_keys:
        payload[key] = data.get(key, None)
    
    # print(payload)    
    # print(data['interior_full'])

    payload["heating"] = payload["cooling"] = None
    if data["interior_full"] != None:
        for dict in data['interior_full']:
            if dict['title'] == "Heating": payload['heating'] = dict['values'][0].split(": ")[1]
            if dict['title'] == "Cooling": payload['cooling'] = dict['values'][0].split(": ")[1]
    return payload


# open our excel doc
# wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
# ws = wb.active

# Code to test our functions
# snapshot_id_parse(zillow_api_call("https://www.zillow.com/homedetails/150-Traincroft-NW-Medford-MA-02155/56277119_zpid/"))
# print(snapshot_id_parse(zillow_api_call("https://www.zillow.com/homedetails/2506-Gordon-Cir-South-Bend-IN-46635/77050198_zpid/")))
# print(zillow_api_call("https://www.zillow.com/homedetails/2506-Gordon-Cir-South-Bend-IN-46635/77050198_zpid/"))

# converts a URL into final zillow data object with the required fields
# test_url = ws.cell(row=4, column=6).value
# test_data = get_zillow_data(get_snapshot_id(test_url))
# print(test_data)




# # loops over the rows in our excel. Bounds are inclusive of first arg, exclusive of second arg. Use the row numbers found in the excel. So to include row 3 you
# # would set your first arg to 3.
# for row in range(155,211):
#     # grab the url for 'row' found in column 6
#     url = ws.cell(row=row, column=6).value

#     # if url does not exist skips to next url
#     if url == "None": continue
#     print(f"Collecting row {row} with url = {url}")

#     # collect the data for the url
#     data = snapshot_id_parse(zillow_api_call(url))

#     # if data collection for the url failed, snapshot_id_parse will return None. Then this conditional will run skipping this row as no data has been collected.
#     if data == None: continue

#     # for each data point, write to a separate column in our excel, if the data point does not exist we write "None"
#     for i,point in enumerate(data.values()):
#         ws.cell(row=row, column=7+i, value=point if point else "None")
    
#     # as changes have been made to the excel in the above code, save the changes
#     wb.save('res-econ_RA_data.xlsx')
    
#     # print(f"data collection for {url} done")
#     sleep(5)

# print("Data collection finished!")