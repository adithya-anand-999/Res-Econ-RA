# This is a code file for working on the custom search engine google API to create a more robust system for URL collections. 
import requests
import openpyxl
import time

from config import API_KEY
from config import CX

def split_str(address): 
    if not isinstance(address, str):
        raise ValueError(f"split_str() did not receive a string")
    cleaned = address.lower().replace(",", "")
    return cleaned.split()


def custom_SE_with_str_check(addr, key=API_KEY, cx=CX):
    # if not addr:
    #     print(f"no address")
    #     return None
    url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}"
    try:
        response = requests.get(url=url).json()
        orig_split = split_str(addr)
        number = orig_split[0]
        name = orig_split[1] # only checking first word of street 
        zip = orig_split[-1] 
        for obj in response['items']: 
            comp_addr = obj['title'].split(" |")[0]
            comp_split = split_str(comp_addr)
            comp_number = comp_split[0]
            comp_name = comp_split[1]
            comp_zip = comp_split[-1]
            # print(f"Checking: {number} == {comp_number}, {name} == {comp_name}, {zip} == {comp_zip}")
            if number in comp_number and name == comp_name and zip == comp_zip: # checking if the target number occurs in range of numbers for this result
                target_url = obj['link']
                # print(f"{addr} → {data}")
                return target_url
    except Exception as err:
        print(f"{err} occurred when processing address: {addr}")
    return "None"


# Older basic API code to collect zillow url for an address
# def custom_SE(addr, key=API_KEY, cx=CX):
#     url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}"
#     try:
#         response = requests.get(url=url).json()
#         data = response['items'][0]['link']
#         print(f"{addr} → {data}")
#         return(data)
#     except Exception as err:
#         print(f"{err} occurred when processing address: {addr}")
#         return None



# open excel
wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
ws = wb.active

# simple tests of the custom_SE function
# addrs = ["6 STANDISH CIR, ANDOVER, MA, 01810", "150 TRAINCROFT ST, MEDFORD, MA, 02155", "16 HARRIS LN, HARVARD, MA, 01451", "24 GREEN VALLEY RD, MEDWAY, MA, 02053"]
# custom_SE_with_str_check(addrs[3])
# split_str(addrs[0])

# converts an address to a zillow URL
test_addr = ws.cell(row=2, column=1).value
test_url = custom_SE_with_str_check(test_addr)
print(test_url)



# code updates excel file res-econ_RA_data.xlsx. Commented out as no reason to run code more than once.
# for row in range(2,205):
#     address = ws.cell(row=row, column=1).value
#     url = custom_SE_with_str_check(address)
#     print(f"{address} → {url}")
#     ws.cell(row=row, column=6, value=url)
#     wb.save('res-econ_RA_data.xlsx')
#     time.sleep(1)
# print("Done updating Excel file.")
