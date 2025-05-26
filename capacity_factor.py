# E

import urllib.request
import json
from config import PV_API_KEY
import openpyxl

"""
The function get_cap_factor takes in latitude and longitude coordinates 
and uses those coordinates to query the PVWatts API and return the 
DC Capacity Factor data.  The loop iterates through the addresses 
coordinates in the address_cord json file, calls get_cap_factor on 
each set of coordinates, the results, alongside the original address, 
latitude, longitude, on which the function was called, are stored in a
dictionary and appended to a list.  The code commented out writes the 
information from the list to an excel document such that each field 
in the dictionary is a column. 
"""

def get_capacity_factor(lat, lon): 
    pv_url = f"https://developer.nrel.gov/api/pvwatts/v8.json?api_key={PV_API_KEY}&lat={lat}&lon={lon}&system_capacity=4&module_type=0&losses=14.08&tilt=20&azimuth=180&array_type=0" 
    response = urllib.request.urlopen(pv_url)
    pv_data = json.loads(response.read())
    return pv_data["outputs"]["capacity_factor"]

# code to test our function 
# test_addr = ws.cell(row=8, column=1).value
# lat = ws.cell(row=8, column=2).value
# lon = ws.cell(row=8, column=3).value
# print(get_capacity_factor(lat, lon))


