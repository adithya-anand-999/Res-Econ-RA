# E

from config import API_KEY
import googlemaps
import openpyxl


def get_coordinates(address, api_key=API_KEY):
    gmaps = googlemaps.Client(key=api_key)
    geocode_result = gmaps.geocode(address)
    if geocode_result:
        location = geocode_result[0]['geometry']['location']
        return (location['lat'], location['lng'])
    return (None)

# open excel
# wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
# ws = wb.active

# code to test our function 
# test_addr = ws.cell(row=8, column=1).value
# print(get_coordinates(test_addr))


# for row in range(2,5):
#     addr = ws.cell(row=row, column=1).value
#     print(get_lat_long_google(addr))





