# This is a code file for working on the custom search engine google API to create a more robust system for URL collections. 
# E

import requests
import openpyxl
import time

from config import API_KEY
from config import CX

"""
The function split_str() is a helper function to assist in string comparisons. 
The function takes in a string address, cleans the text, splits the address by spaces, 
and returns a list in which each word from the address is an element in the list. 
Ex. 
    input: '175 Fifth Avenue, New York, NY 10010'  
    output: ['175', 'fifth', 'avenue', 'new', 'york', 'ny', '10010']
"""

def split_str(address): # address = string, written address 
    if not isinstance(address, str): # checks if the address is not a string 
        raise ValueError(f"split_str() did not receive a string") 
    cleaned = address.lower().replace(",", "") # if the address is a string, the text is converted to lower case and commas are removed
    return cleaned.split() # the cleaned address is split by spaces, a list of strings with each word is returned 

"""
The function get_zillow_url() searches specifically for Zillow urls for a 
particular address through the use of a Google Custom Search Engine. 
The function takes in a written address, the Google API Key, and the 
Google Custom Search Engine Key.  The parameters are used to build a custom
Google search for Zillow urls and returns the url that matches information from 
the input address.  The Google API Key and Google Custom Search Engine Key 
are defaulted to the keys designated in the config.py file. 
"""

def get_zillow_url(addr, key=API_KEY, cx=CX): # addr = address to search, key = Google API Key, cx = Google Custom Search Engine Key 
    # if not addr:
    #     print(f"no address")
    #     return None
    url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}" # builds the custom search request url using each parameter 
    try:
        response = requests.get(url=url).json() # sends a GET request to the custom search url, parses the response as a json
        orig_split = split_str(addr) # calls the helper function to clean and split the address 
        number = orig_split[0] # from the split address, gets the first value (address number)
        name = orig_split[1] # from the split address, gets the second value (street name - only checks first word of street name if name is longer than one word)
        zip = orig_split[-1] # from the split address, gets the last value (zip code) 
        for obj in response['items']:  # loops through each search result 
            comp_addr = obj['title'].split(" |")[0] # extracts the address portion of each search result, based on Google search result formatting
            comp_split = split_str(comp_addr) # calls the helper function to clean and split the address 
            comp_number = comp_split[0] # from the split address, gets the first value (address number)
            comp_name = comp_split[1] # from the split address, gets the second value (street name - only checks first word of street name if name is longer than one word)
            comp_zip = comp_split[-1] # from the split address, gets the last value (zip code) 
            # print(f"Checking: {number} == {comp_number}, {name} == {comp_name}, {zip} == {comp_zip}")
            if number in comp_number and name == comp_name and zip == comp_zip: 
            # the above line checks if the input address number appears in result's number, checks that street name and zip code match exactly between input and result 
                target_url = obj['link'] 
                # print(f"{addr} → {data}")
                return target_url # if the result passes the check, the Zillow url is accessed and returned
    except Exception as err: # if any error occurs, the below error message is displayed
        print(f"{err} occurred when processing address: {addr}")
    return "None" # if a match is not found or an error occurs, None is returned 


# Older basic API code to collect zillow url for an address
# def custom_SE(addr, key=API_KEY, cx=CX):
#     url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}"
#     try:
#         response = requests.get(url=url).json()
#         data = response['items'][0]['link']
#         print(f"{addr} → {data}")
#         return(data)
#     except Exception as err:
#         print(f"{err} occurred when processing address: {addr}")
#         return None



# open excel
# wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
# ws = wb.active

# simple tests of the custom_SE function
# addrs = ["6 STANDISH CIR, ANDOVER, MA, 01810", "150 TRAINCROFT ST, MEDFORD, MA, 02155", "16 HARRIS LN, HARVARD, MA, 01451", "24 GREEN VALLEY RD, MEDWAY, MA, 02053"]
# custom_SE_with_str_check(addrs[3])
# split_str(addrs[0])

# converts an address to a zillow URL
# test_addr = ws.cell(row=2, column=1).value
# test_url = get_zillow_url(test_addr)
# print(test_url)



# code updates excel file res-econ_RA_data.xlsx. Commented out as no reason to run code more than once.
# for row in range(2,205):
#     address = ws.cell(row=row, column=1).value
#     url = custom_SE_with_str_check(address)
#     print(f"{address} → {url}")
#     ws.cell(row=row, column=6, value=url)
#     wb.save('res-econ_RA_data.xlsx')
#     time.sleep(1)
# print("Done updating Excel file.")
