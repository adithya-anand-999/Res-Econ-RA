from googlesearch import search
import openpyxl
import time

def get_zillow_url(address):
    query = f'site:zillow.com "{address}"'
    for url in search(query, num_results=5):
        if "zillow.com/homedetails" in url:
            return url.split('?')[0]  # Remove any tracking parameters
    return None

# Load the Excel workbook and select the active worksheet.
wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
ws = wb.active

# Iterate over the rows starting from the second (assuming row 1 has headers).
for row in range(90, ws.max_row + 1):
    address = ws.cell(row=row, column=1).value  # Assuming addresses are in column A.
    # if not address:
    #     continue  # Skip rows without an address.
    
    print(f"Searching for: {address}")
    url = get_zillow_url(address)
    print("→", url)
    
    # Write the URL to column B (or any column of your choice).
    ws.cell(row=row, column=6, value=url)
    
    # Save the workbook after processing each address.
    wb.save('res-econ_RA_data.xlsx')
    
    # Sleep for one second to prevent triggering rate limits.
    time.sleep(1)

print("Done updating Excel file.")
