import requests
import openpyxl
import time


from config import API_KEY


def solar_api(lat, lng):
    url = "https://solar.googleapis.com/v1/buildingInsights:findClosest"
    params = {
        "location.latitude": f"{lat:.5f}",
        "location.longitude": f"{lng:.5f}",
        "key": API_KEY,
    }

    try:
        response = requests.get(url, params=params)
        content = response.json()
        # print(content)
        return round(content['solarPotential']['maxArrayAreaMeters2'] * 10.7639,4)
    except Exception as e:
        print(f"an error occurred {e}")

# individual testing of code
# lat = 42.3202535
# long = -71.3471081
# print(solar_api(lat,long))


# excel writing of code
wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
ws = wb.active

for row in range(4,55):
    lat = ws.cell(row=row, column=2).value
    long = ws.cell(row=row, column=3).value
    data = solar_api(lat,long)
    print(f"{lat,long} → {data}")
    ws.cell(row=row, column=6, value=data)
    wb.save('res-econ_RA_data.xlsx')
    time.sleep(1)
print("Done updating Excel file.")