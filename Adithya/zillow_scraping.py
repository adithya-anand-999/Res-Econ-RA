from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
import openpyxl
from bs4 import BeautifulSoup

wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
ws = wb.active

options = Options()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                     "AppleWebKit/537.36 (KHTML, like Gecko) "
                     "Chrome/122.0.0.0 Safari/537.36")

driver = webdriver.Chrome(options=options)

for row in range(2,ws.max_row+1):
    try:
        url = ws.cell(row=row, column=7).value
        driver.get(url)
        # time.sleep(5)
        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1[data-testid='home-details-summary-headline']")))
        except Exception as e:
            print(f"Timeout waiting for Zillow content: {e}")
            break
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        print(soup.prettify())
        input("Press Enter to continue to next URL...")
    except Exception as e:
        print(f"Request failed for {url}: {e}")
        break
    time.sleep(30)
driver.quit()