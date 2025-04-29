import os
import openpyxl
import time

import nest_asyncio
nest_asyncio.apply()

from requests_html import AsyncHTMLSession
from bs4 import BeautifulSoup
import asyncio
# import pyppeteer.errors
# os.environ["PYPPETEER_EXECUTABLE_PATH"] = r" C:\Program Files\Google\Chrome\Application\chrome.exe"

asession = AsyncHTMLSession()

async def scrape_roof_space(addr_num,lat,lon):
    url = f"https://sunroof.withgoogle.com/building/{str(round(float(lat),4))}/{str(round(float(lon),4))}/#?f=buy"

    try:
        response = await asession.get(url)
        await response.html.arender(timeout=20)
        soup = BeautifulSoup(response.html.html, "html.parser")
        # print(soup)

        inp = soup.find("input", attrs={"aria-label": True})
        if not inp:
            inp = soup.find("input", attrs={"placeholder": True})

        if inp:
            address = inp.get("aria-label") or inp.get("placeholder")
            # print("Address:", address)
            # print(address.split(" ")[0])
            if addr_num != address.split(" ")[0]: return "None"
        else:
            print("Couldn’t find the address input in the page.")
            return "None"
        


        upper_li = soup.find("li", attrs={"ng-if": "$ctrl.derivedValues.getMaxArrayAreaSqft()"})
        if upper_li:
            data_div = upper_li.find("div", class_="panel-fact-text md-body")
            if data_div:
                val = data_div.get_text(strip=True).split()[0]
                # print(val)
                return(val)
            else:
                print("Inner data div not found.")
                return("Inner data div not found.")
        else:
            print("Upper div with specified ng-if not found.")
            return("Upper div with specified ng-if not found.")
    except Exception as e:
        print(f"Uncaught error {e} occurred")
        return("None")

# asyncio.run(scrape_roof_space("6","42.6576", "-71.1582"))


# open excel file
wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
ws = wb.active

# code to test our function
test_url = ws.cell(row=4, column=6).value
test_data = snapshot_id_parse(zillow_api_call(test_url))
print(test_data)


# writing results to excel
# for row in range(150,211):
#     addr = ws.cell(row=row, column=1).value
#     # print(addr_num)
#     lat = ws.cell(row=row, column=2).value
#     long = ws.cell(row=row, column=3).value
#     data = asyncio.run(scrape_roof_space(addr.split(" ")[0],lat,long))
#     print(f"{addr} → {data}")
#     ws.cell(row=row, column=6, value=data)
#     wb.save('res-econ_RA_data.xlsx')
#     time.sleep(1)
# print("Done updating Excel file.")