# This is a code file for working on the custom search engine google API to create a more robust system for URL collections. 
import requests
import openpyxl
import time

from config import API_KEY
from config import CX

def split_str(address): 
    lower = address.lower()
    cleaned = lower.replace(",", "")
    split = cleaned.split()
    return split

def custom_SE_with_str_check(addr, key=API_KEY, cx=CX):
    url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}"
    try:
        response = requests.get(url=url).json()
        orig_split = split_str(addr)
        number = orig_split[0]
        name = orig_split[1] # maybe an issue if street name longer than 1 word
        zip = orig_split[-1] # used -1 in case of different length arrays (e.g. due to additional words in address)
        for obj in response['items']: 
            comp_addr = obj['title'].split(" |")[0]
            comp_split = split_str(comp_addr)
            comp_number = comp_split[0]
            comp_name = comp_split[1]
            comp_zip = comp_split[-1]
            print(f"Checking: {number} == {comp_number}, {name} == {comp_name}, {zip} == {comp_zip}")
            if number == comp_number and name == comp_name and zip == comp_zip: 
                data = obj['link']
                print(f"{addr} → {data}")
                return data
    except Exception as err:
        print(f"{err} occurred when processing address: {addr}")
        return None

addrs = ["6 STANDISH CIR, ANDOVER, MA, 01810", "150 TRAINCROFT ST, MEDFORD, MA, 02155", "16 HARRIS LN, HARVARD, MA, 01451", "24 GREEN VALLEY RD, MEDWAY, MA, 02053"]
custom_SE_with_str_check(addrs[3])
# split_str(addrs[0])

# code gets correct url for 16 Harris (address with inital issue) and with 2-word street names 

# def custom_SE(addr, key=API_KEY, cx=CX):
#     url=f"https://www.googleapis.com/customsearch/v1?key={key}&cx={cx}&q={addr}"
#     try:
#         response = requests.get(url=url).json()
#         data = response['items'][0]['link']
#         print(f"{addr} → {data}")
#         return(data)
#     except Exception as err:
#         print(f"{err} occurred when processing address: {addr}")
#         return None


# simple tests of the custom_SE function

# code updates excel file res-econ_RA_data.xlsx. Commented out as no reason to run code more than once.
# wb = openpyxl.load_workbook('./res-econ_RA_data.xlsx')
# ws = wb.active

# for row in range(1,ws.max_row+1):
#     address = ws.cell(row=row, column=1).value
#     url = custom_SE(address)
#     ws.cell(row=row, column=7, value=url)
#     wb.save('res-econ_RA_data.xlsx')
#     time.sleep(1)

# print("Done updating Excel file.")
